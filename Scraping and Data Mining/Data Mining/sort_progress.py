from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting import CellIsRule
import sys, time, pprint

wb1 = load_workbook('reddit_scraper_test.xlsx')
wb2 = load_workbook('reddit_scraper_test2.xlsx')
wb_change = load_workbook('change_in_test_rank.xlsx')

ws1 = wb1.active
ws2 = wb2.active
ws_change = wb_change.active

blackFill = PatternFill(start_color='00000000',
                        end_color='00000000',
                        fill_type='solid')

whiteFont = Font(color='FFFFFFFF')
                                     


rows2 = ws2.rows
rows_list = []
for r2 in rows2:
    for cell in r2:
        rows_list.append(cell)


rows1 = ws1.rows
rows1_list = []
for r1 in rows1:
    rows1_list.append(r1)

start = time.time()
track_row = 2
for row in range(2, len(rows1_list)):
    value = ws1.cell(row = row, column = 7).value
    
    for cell in rows_list:
        if value in str(cell.value):
            rank1 = row - 1
            rank2 = cell.row - 1
            _ = ws_change.cell(row=track_row, column=1, value=value)
            _ = ws_change.cell(row=track_row, column=2, value=rank1)
            _ = ws_change.cell(row=track_row, column=3, value=rank2)

            if rank1 == rank2:
                _ = ws_change.cell(row=track_row, column=4, value='-')

            elif rank1 > rank2:
                rank_diff = rank1 - rank2
                _ = ws_change.cell(row=track_row, column=4, value=('+%s' %rank_diff))

            elif rank2 < rank1:
                rank_diff = rank2 - rank1
                _ = ws_change.cell(row=track_row, column=4, value=('-%s' %rank_diff))

            else:
                print('what the fuck did I miss... rank1 = %s, rank2 = %s' %(rank1, rank2))
            
            track_row += 1

cols_change = ws_change.columns
col3 = []
for col in cols_change:
    col3.append(col)
col3 = col3[0]
col_list = []
for cell in col3:
    col_list.append(cell.value)

    
track_new_row = track_row
track_row = 2
for row in range(2, len(rows1_list)):
    value = ws1.cell(row=row, column=7).value

    if value not in col_list:
        _ = ws_change.cell(row=track_row, column=5, value=('| '+str(track_new_row)))
        _ = ws_change.cell(row=track_row, column=6, value=value)
        track_new_row += 1
        track_row += 1

col = ws_change.columns[4]
for cell in col:
    cell.font = whiteFont
    cell.fill = blackFill

timediff = time.time() - start
duration = time.strftime('%M:%S', time.localtime(timediff))
wb_change.save('change_in_test_rank_result.xlsx')
print('Done.\nTotal time elapsed: %s' %(duration))
