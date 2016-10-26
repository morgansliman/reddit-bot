from openpyxl import Workbook, load_workbook
import praw, datetime, time, sys

r = praw.Reddit(user_agent = 'funny scraper v1')
subreddit = r.get_subreddit('funny')
wb = load_workbook('reddit_scraper.xlsx')
ws = wb['funny']

# Column A = 'Title'            # Post Title
# Column B = 'Score as of >'    # Post score
# Column C = 'Time gathered'    # Current time when score recorded
# Column D = 'Time created'     # time submission was created
# Column E = 'by author >'      # author of post(s)
# Column F = 'within range:'    # scrape range
# Column G = 'Submission ID'    # submission ID
limit = None
row = 2
t = time.time()
top = subreddit.get_top_from_day(limit=limit)
for post in top:
 #   print('writing post #%s' %(row-1))
    column = 1
    title = post.title
    score = post.score
    now = datetime.datetime.now()
    author = str(post.author)
    rng = limit
    sub_id = post.id
    time_created = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(post.created))

    var_list = [title, score, now, time_created, author, rng, sub_id]
    while column < 8:
        try:
            _ = ws.cell(row=row, column=column, value=var_list[column-1])
        except:
            print('list index %s out of range %s' %(column, len(var_list)))
            sys.exit(1)
        column += 1

    row += 1
timediff = time.time() - t
duration = time.strftime('%M:%S', time.localtime(timediff))
print('Done.\nTotal time elapsed: %s' %(duration))
wb.save('reddit_scraper_test2.xlsx')
